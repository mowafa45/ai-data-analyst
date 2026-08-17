"""
Export Router — /api/export
POST /pdf    — generate PDF report
POST /excel  — generate Excel summary
POST /csv    — raw CSV download
"""
import io
from datetime import datetime

import numpy as np
import structlog
from fastapi import APIRouter, HTTPException, status
from fastapi.responses import StreamingResponse

from models.schemas import ExportRequest
from services.data_service import load_dataframe, load_meta
from services.insight_service import generate_dashboard

log = structlog.get_logger()
router = APIRouter()


@router.post("/pdf")
async def export_pdf(req: ExportRequest):
    """Generate a PDF report with KPIs, charts, and insights."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )

        df = await load_dataframe(req.session_id)
        meta = await load_meta(req.session_id)
        dashboard = await generate_dashboard(req.session_id)

        if df is None or meta is None:
            raise ValueError("Session not found.")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()

        primary = colors.HexColor("#6366f1")
        dark = colors.HexColor("#0f0f11")
        muted = colors.HexColor("#6b7280")

        title_style = ParagraphStyle("title", parent=styles["Title"], textColor=primary, fontSize=24, spaceAfter=6)
        h2_style = ParagraphStyle("h2", parent=styles["Heading2"], textColor=dark, fontSize=14, spaceBefore=16, spaceAfter=8)
        body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14, spaceAfter=6)
        muted_style = ParagraphStyle("muted", parent=styles["Normal"], fontSize=9, textColor=muted)

        story = []

        # Header
        story.append(Paragraph("AI Data Analyst Report", title_style))
        story.append(Paragraph(f"Dataset: {meta.filename} · Generated {datetime.utcnow().strftime('%B %d, %Y %H:%M UTC')}", muted_style))
        story.append(Paragraph(f"{meta.row_count:,} rows · {meta.col_count} columns · {meta.duplicates_removed} duplicates removed", muted_style))
        story.append(HRFlowable(width="100%", thickness=1, color=primary, spaceAfter=12))

        # KPIs
        story.append(Paragraph("Key Performance Indicators", h2_style))
        kpi_data = [["Metric", "Value", "Trend"]]
        for kpi in dashboard.kpis:
            arrow = "▲" if kpi.trend == "up" else "▼" if kpi.trend == "down" else "—"
            delta = f"{arrow} {abs(kpi.delta_pct or 0):.1f}%" if kpi.delta_pct is not None else "—"
            kpi_data.append([kpi.label, kpi.value, delta])

        kpi_table = Table(kpi_data, colWidths=[8*cm, 5*cm, 4*cm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), primary),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f9fafb"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.5*cm))

        # Insights
        if req.include_insights:
            story.append(Paragraph("AI-Generated Insights", h2_style))
            for ins in dashboard.insights:
                story.append(Paragraph(f"<b>{ins.emoji} {ins.headline}</b>", body_style))
                story.append(Paragraph(ins.detail, body_style))
                story.append(Paragraph(
                    f"Columns used: {', '.join(ins.columns_used)} · {ins.row_count_analyzed:,} rows · Confidence: {ins.confidence*100:.0f}%",
                    muted_style
                ))
                story.append(Spacer(1, 0.3*cm))

        # Recommendations
        story.append(Paragraph("Recommendations", h2_style))
        for i, rec in enumerate(dashboard.recommendations, 1):
            story.append(Paragraph(f"{i}. {rec}", body_style))

        # Data sample
        story.append(Paragraph("Data Sample (first 10 rows)", h2_style))
        numeric_cols = [c.name for c in meta.columns if c.dtype == "numeric"][:5]
        cat_cols = [c.name for c in meta.columns if c.dtype == "categorical"][:3]
        show_cols = (cat_cols + numeric_cols)[:6]
        if show_cols:
            sample = df[show_cols].head(10).replace({np.nan: None})
            tbl_data = [show_cols] + [[str(v) if v is not None else "" for v in row] for row in sample.values.tolist()]
            tbl = Table(tbl_data)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(tbl)

        doc.build(story)
        buf.seek(0)

        filename = f"ai_analyst_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        log.error("PDF export failed", error=str(e))
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


@router.post("/excel")
async def export_excel(req: ExportRequest):
    """Generate an Excel workbook with multiple analysis sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers

        df = await load_dataframe(req.session_id)
        meta = await load_meta(req.session_id)
        dashboard = await generate_dashboard(req.session_id)

        if df is None or meta is None:
            raise ValueError("Session not found.")

        wb = openpyxl.Workbook()

        # ── Sheet 1: Dashboard ───────────────────────────────────────────────
        ws = wb.active
        ws.title = "Dashboard"
        purple = "6366F1"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor=purple)

        ws["A1"] = "AI Data Analyst — Dashboard"
        ws["A1"].font = Font(bold=True, size=16, color=purple)
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A3"] = f"Dataset: {meta.filename}  |  {meta.row_count:,} rows  |  {meta.col_count} columns"

        ws.append([])
        ws.append(["Metric", "Value", "Trend", "Delta"])
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill

        for kpi in dashboard.kpis:
            ws.append([kpi.label, kpi.value, kpi.trend.upper(), f"{kpi.delta_pct:+.1f}%" if kpi.delta_pct else "—"])

        ws.append([])
        ws.append(["AI Insights"])
        ws[ws.max_row][0].font = Font(bold=True, size=12)
        for ins in dashboard.insights:
            ws.append([f"{ins.emoji} {ins.headline}", ins.detail, f"Confidence: {ins.confidence*100:.0f}%"])

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12

        # ── Sheet 2: Raw Data ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Data")
        headers = list(df.columns)
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in df.replace({np.nan: None}).head(5000).itertuples(index=False):
            ws2.append(list(row))

        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 15

        # ── Sheet 3: Column Stats ────────────────────────────────────────────
        ws3 = wb.create_sheet("Column Stats")
        ws3.append(["Column", "Type", "Null Count", "Null %", "Unique Values", "Sample"])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill

        for col in meta.columns:
            ws3.append([
                col.name, col.dtype, col.null_count, f"{col.null_pct:.1f}%",
                col.unique_count, ", ".join(str(v) for v in col.sample_values[:3]),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"ai_analyst_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        log.error("Excel export failed", error=str(e))
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")


@router.post("/csv")
async def export_csv(req: ExportRequest):
    """Export the cleaned dataset as CSV."""
    df = await load_dataframe(req.session_id)
    if df is None:
        raise HTTPException(status_code=404, detail="Session not found.")

    buf = io.StringIO()
    df.to_csv(buf, index=False)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="cleaned_data.csv"'},
    )
